#  Đây là để tài luận văn Mobile robot và chatbot được kết hợp với nhau tạo nên mô hình MOBILE CHATBOT
#  Đề tài được thực hiện tại Đại học kỹ thuật công nghệ cần thơ CTUET trong bói cảnh đại dịch Covid19
#  chúng tôi gặp rất nhiều khó khăn trong lúc thực hiện xin kể ra đây: Hư màn hình laptop, cháy vi điều khiển, thường overnight, kinh phí cho kinh nghiệm,...

# =====================================================//=======================================
#  Giảng viên hướng dẫn: Th.S Nguyễn Lê Thế Duy
#  SVTH: Nguyễn Minh Phước  & Lê Trung Kiên
#  Lớp CDT0117
#  Ngành công nghệ kỹ thuật cơ điện tử - Khoa kỹ thuật cơ khí
#  hoàn thành đề tài: 06/2021

# =================================================//==================================
#  import các thư viện cần thiết // lưu ý các biên bản phải phù hợp
# python 3.7.9
# tensorflow 1.5


#Library Chung
import time
import numpy as np
import serial.tools.list_ports
#Library Kien
import random
import json
import openpyxl
import torch
import pyttsx3
import speech_recognition
import webbrowser
import os
import playsound

from gtts import gTTS
from googletrans import Translator
from tkinter import *
from tkinter import messagebox
from datetime import datetime
from model import NeuralNet
from nltk_utils import bag_of_words, tokenize

#Library Phuoc


import tensorflow as tf
import cv2

from utils import label_map_util
from utils import visualization_utils_color as vis_util
from imutils.video import FPS
from imutils.video import WebcamVideoStream

# Khai Bao Arduino
#
ports = list(serial.tools.list_ports.comports())
for p in ports:

    if "CH340" in p.description:
        a=p.description.split()
        print("This is an Arduino!")
        b=a[len(a)-1].replace("(", "").replace(")", "")
        # ArduinoSerial = serial.Serial(port=str(b),baudrate=9600,timeout=0.2)
        print(p)
    if "Arduino" in p.description:
        a=p.description.split()
        print("This is an Arduino!")
        b=a[len(a)-1].replace("(", "").replace(")", "")
        # ArduinoSerial = serial.Serial(port=b,baudrate=9600,timeout=0.2)
        print(p)

ArduinoSerial = serial.Serial(port=b,baudrate=9600,timeout=0.2)
# print(ArduinoSerial)
# ArduinoSerial = serial.Serial('Com5', 9600,timeout=0.2)
time.sleep(1)

#Khai Bao Kien

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

with open('intents.json', 'r') as json_data:
    intents = json.load(json_data)

FILE = "data.pth"
data = torch.load(FILE)
input_size = data["input_size"]
hidden_size = data["hidden_size"]
output_size = data["output_size"]
all_words = data['all_words']
tags = data['tags']
model_state = data["model_state"]

model = NeuralNet(input_size, hidden_size, output_size).to(device)
model.load_state_dict(model_state)
model.eval()

robot_mouth = pyttsx3.init()
robot_ear = speech_recognition.Recognizer()

now = datetime.now()
bot_name = "KP"
cau_hoi_them = "Do you need any more help"
print("Let's Start")
print(bot_name, ": Please wait a few minutes to run the program")

lg = ["none"]
bt = ["none"]
quataiCDT = []
quataiTDH = []
#Khai Bao Phuoc
cam = WebcamVideoStream(1).start()
fps = FPS().start()
    # Path to frozen detection graph. This is the actual model that is used for the object detection.
PATH_TO_CKPT = './model/frozen_inference_graph.pb'

    # List of the strings that is used to add correct label for each box.
PATH_TO_LABELS = './proto/label_map.pbtxt'
NUM_CLASSES = 1

    # Loading label map
label_map = label_map_util.load_labelmap(PATH_TO_LABELS)
categories = label_map_util.convert_label_map_to_categories(label_map, max_num_classes=NUM_CLASSES, use_display_name=True)
category_index = label_map_util.create_category_index(categories)


# Send Uart Arduino
    #Open CĐT
def Open_A():
    ArduinoSerial.write(('11').encode())

    #Open TĐH
def Open_B():
    ArduinoSerial.write(('12').encode())

def rst():
    ArduinoSerial.write(('14').encode())

    #open door
def door():
    ArduinoSerial.write(('3').encode())
    time.sleep(0.2)
    ArduinoSerial.write(('4').encode())
    time.sleep(1)
    ArduinoSerial.write(('3').encode())
    time.sleep(0.2)
    ArduinoSerial.write(('4').encode())
    print("done!")

    # Robot Forward
def Forward():
    ArduinoSerial.write(('6').encode())

    # Robot Reverse
def Reverse():
    ArduinoSerial.write(('7').encode())

    # Robot Turn Right
def Turnright():
    ArduinoSerial.write(('8').encode())

    # Robot Turn Left
def Turnleft():
    ArduinoSerial.write(('5').encode())

     # Robot Stop
def Stop():
    ArduinoSerial.write(('0').encode())

######## Def Kien

def open_CDT():
    quataiCDT.append(1 + len(quataiCDT))
    # print("len",len(quataiTDH))
    if len(quataiCDT) > 100:
        for i in range(len(bt)):
            bt.pop()
            # print("i", i)
        bt.append("cdt")

        if lg[0] == "vn":
            var_bot.set("Xin Lỗi. Ngăn CĐT Đã Đầy")
            window.update()
            time.sleep(2)
            chatbot_vn()
        else:
            var_bot.set("Sorry. Mechatronic Is Full")
            window.update()
            time.sleep(2)
            chatbot_en()
    else:
        for i in range(len(bt)):
            bt.pop()
            # print("i", i)
        bt.append("cdt")
        Open_A()

        if lg[0] == "vn":
            var_bot.set("Đang mở ngăn chứa của Cơ Điện Tử, Đợi một lát")
            window.update()
        else:
            var_bot.set("Opening Mechatronic. Please wait few second")
            window.update()

        path = 'D:\login_picture\CDT'
        now = datetime.now()
        d = now.strftime("%d_%m_%Y_%H_%M_%S")
        frame = cam.read()
        filename = "CDT-" + d + ".jpg"
        cv2.imwrite(os.path.join(path, filename), frame)

        link = 'HYPERLINK' + '(' + '"' + 'D:\login_picture\CDT-' + d + ".jpg" + '")'
        # print("l", link)

        wb = openpyxl.load_workbook('Data_Management.xlsx')
        now = datetime.now()
        sh1 = wb['Data_times_open_door']
        row = sh1.max_row + 1
        stt = row - 1
        sh1.cell(row=row, column=1, value=stt)
        sh1.cell(row=row, column=2, value=now.strftime("%d"))
        sh1.cell(row=row, column=3, value="Tháng " + now.strftime("%m"))
        sh1.cell(row=row, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
        sh1.cell(row=row, column=5, value=link)
        sh1.cell(row=row, column=6, value="Khoa CĐT")
        wb.save("Data_Management.xlsx")

        if lg[0] == "vn":
            print("Đang mở ngăn chứa của Cơ Điện Tử")
            time.sleep(12)
            rst()
            var_user.set(' ')
            var_bot.set("Hoàng Thành!")
            window.update()
            print("Hoàng Thành!")
            time.sleep(2)
            chatbot_vn()
        else:
            print("Open CDT")
            time.sleep(12)
            rst()
            var_user.set(' ')
            var_bot.set("Done!")
            window.update()
            time.sleep(2)
            print("Done")
            chatbot_en()

def open_TDH(): ##Open TDH
    quataiTDH.append(1 + len(quataiTDH))
    # print("len",len(quataiTDH))
    if len(quataiTDH) > 100:
        for i in range(len(bt)):
            bt.pop()
            # print("i", i)
        bt.append("tdh")

        if lg[0] == "vn":
            var_bot.set("Xin Lỗi. Ngăn TĐH Đã Đầy")
            window.update()
            time.sleep(2)
            chatbot_vn()
        else:
            var_bot.set("Sorry. Automation Is Full")
            window.update()
            time.sleep(2)
            chatbot_en()
    else:
        for i in range(len(bt)):
            bt.pop()
            # print("i", i)
        bt.append("tdh")
        Open_B()

        if lg[0] == "vn":
            var_bot.set("Đang mở ngăn chứa của khoa Tự Động Hóa, Đợi một lát")
            window.update()
        else:
            var_bot.set("Opening Automation. Please wait few second")
            window.update()

        path = 'D:\login_picture\TDH'
        now = datetime.now()
        d = now.strftime("%d_%m_%Y_%H_%M_%S")
        frame = cam.read()
        filename = "TDH-"+ d +".jpg"
        cv2.imwrite(os.path.join(path, filename), frame)

        link = 'HYPERLINK' + '(' + '"' + 'D:\login_picture\Admin_access_to_system\TDH-'+ d +".jpg" + '")'
        # print("l", link)


        wb = openpyxl.load_workbook('Data_Management.xlsx')
        now = datetime.now()
        sh1 = wb['Data_times_open_door']
        row = sh1.max_row + 1
        stt = row - 1
        sh1.cell(row=row, column=1, value=stt)
        sh1.cell(row=row, column=2, value=now.strftime("%d"))
        sh1.cell(row=row, column=3, value="Tháng " + now.strftime("%m"))
        sh1.cell(row=row, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
        sh1.cell(row=row, column=5, value=link)
        sh1.cell(row=row, column=6, value="Khoa TĐH")
        wb.save("Data_Management.xlsx")

        if lg[0]=="vn":
            print("Đang mở ngăn chứa của khoa Tự Động Hóa")
            time.sleep(12)
            rst()
            var_user.set(' ')
            var_bot.set("Hoàng Thành!")
            window.update()
            print("Hoàng Thành!")
            time.sleep(2)
            chatbot_vn()
        else:
            print("Open TDH")
            time.sleep(12)
            rst()
            var_user.set(' ')
            var_bot.set("Done!")
            window.update()
            time.sleep(2)
            print("Done")
            chatbot_en()

def chatbot_en(): #Chatbot english
    voices = robot_mouth.getProperty('voices')
    robot_mouth.setProperty('voice', voices[0].id)
    robot_mouth.setProperty('rate', 120)

    if bt[0] == "none":
        var_user.set("   ")
        var_bot.set("Nice To Meet You")
        window.update()
        robot_mouth.say("Nice To Meet You")
        robot_mouth.runAndWait()
        time.sleep(1)
        var_bot.set("Please choose vietnamese or english language")
        window.update()

        robot_mouth.say("please choose vietnamese or english language")
        robot_mouth.runAndWait()
    else:
        var_user.set("   ")
        var_bot.set("Nice To Meet You")
        window.update()

        robot_mouth.say("Nice To Meet You")
        robot_mouth.runAndWait()

    dem1=0
    while True:
        with speech_recognition.Microphone() as mic:
            print(bot_name, ":I'm Listening")
            var_bot.set(bot_name + " :I'm Listening")
            window.update()
            audio = robot_ear.listen(mic, timeout=4, phrase_time_limit=4)
        try:
            sentence = robot_ear.recognize_google(audio)
        except:
            sentence = ""

        print("You: " + sentence)
        Yousen =f"{sentence}"
        var_user.set("You: " +Yousen)
        window.update()
        #print(Err)

        sentence = tokenize(sentence)
        X = bag_of_words(sentence, all_words)
        X = X.reshape(1, X.shape[0])
        X = torch.from_numpy(X).to(device)

        output = model(X)
        _, predicted = torch.max(output, dim=1)
        tag = tags[predicted.item()]
        probs = torch.softmax(output, dim=1)
        prob = probs[0][predicted.item()]
        print(prob)
        if prob.item() > 0.75:
            for intent in intents['intents']:
                if tag == intent["tag"]:
                    fb = f" {random.choice(intent['responses'])}"
                    print(bot_name, ":" + fb)

                    var_bot.set(bot_name+":"+fb)
                    window.update()

                    robot_mouth.say(fb)
                    robot_mouth.say(cau_hoi_them)
                    robot_mouth.runAndWait()
                    print(tag)

            if tag == "goodbye":
                detect_red()
                print("hh")
                break
            elif tag == "send faculty Mechatronic":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("cdt")
                open_CDT()
                break
            elif tag == "send faculty Automation":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("tdh")
                open_TDH()
                break
            elif tag == "laguane Vn":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("vn")
                chatbot_vn()
                print("VN")
                break
            elif tag == "open youtube":
                webbrowser.open("youtube.com")
                var_bot.set("Please choose vietnamese or english language")
                window.update()
                break
            elif tag == "Introduction Faculty":
                # webbrowser.open("https://www.youtube.com/watch?v=KjiNHl-LTOc")
                var_bot.set("Please watch the video to know more")
                window.update()
                time.sleep(10)
                # break
            elif tag == "University introduction":
                # webbrowser.open("https://www.youtube.com/watch?v=iEoO2Jw59vs&t=419s")
                var_bot.set("Please watch the video to know more")
                window.update()
                break
        else:
            Err = f"Err{sentence}"
            print("E:", Err)
            if Err == "Err[]":
                dem1 = dem1 + 1
                var_user.set("You: Hmm")
                var_bot.set(bot_name + "Hmm")
                window.update()

            if Err != "Err[]":
                dem1 = 0
                print("luud ta")
                wb = openpyxl.load_workbook('Data_Management.xlsx')
                sh2 = wb['Data_Err']
                row2 = sh2.max_row + 1
                now = datetime.now()
                stt = row2 - 1
                sh2.cell(row=row2, column=1, value=stt)
                sh2.cell(row=row2, column=2, value=now.strftime("%d"))
                sh2.cell(row=row2, column=3, value="Tháng " + now.strftime("%m"))
                sh2.cell(row=row2, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
                sh2.cell(row=row2, column=5, value=Err)
                wb.save("Data_Management.xlsx")

                var_bot.set(bot_name + " :I do not understand, Can you speak again")
                window.update()

                robot_mouth.say("I do not understand, can you Ask some questions like")
                robot_mouth.runAndWait()
                time.sleep(0.5)

                chd = []
                for intent in intents['intents']:
                    chd.append(f"-{random.choice(intent['patterns'])}")
                # cauhd=(chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)])
                cauhd = ["Ask some questions like:", "\n", chd[0], "\n", chd[1], "\n", chd[2], "\n", chd[3],
                         "\n", chd[4], "\n", chd[5], "\n", chd[6], "\n", chd[7], "\n", chd[8], "\n", chd[9],
                         "\n", chd[10]]
                var_user.set("-")
                var_bot.set(cauhd)
                window.update()
                time.sleep(7)
                print(f"{bot_name} : I do not understand...")

        if dem1 > 3:
            var_bot.set("Bye, see you later")
            var_user.set("----")
            window.update()
            robot_mouth.say("Bye, see you later")
            robot_mouth.runAndWait()
            detect_red()
            break

def chatbot_vn():
    if bt[0] == "none":
        var_user.set("   ")
        var_bot.set("Rất vui được gặp bạn")
        window.update()

        output = gTTS("Rất vui được gặp bạn", lang="vi", slow=False)
        output.save("output.mp3")
        playsound.playsound('output.mp3', True)
        os.remove("output.mp3")

        time.sleep(1)
        var_bot.set("Chọn ngôn ngữ tiếng Anh hoặc tiếng Việt")
        window.update()

        output = gTTS("Chọn ngôn ngữ tiếng Anh hoặc tiếng Việt", lang="vi", slow=False)
        output.save("output.mp3")
        playsound.playsound('output.mp3', True)
        os.remove("output.mp3")

    else:
        var_user.set("   ")
        var_bot.set("Rất vui được gặp bạn")
        window.update()

        output = gTTS("Rất vui được gặp bạn", lang="vi", slow=False)
        output.save("output.mp3")
        playsound.playsound('output.mp3', True)
        os.remove("output.mp3")

    r = speech_recognition.Recognizer()
    dem1=0
    while True:

        with speech_recognition.Microphone() as source:
            print("Tôi Đang Nghe: ")
            var_bot.set(bot_name + " :Tôi Đang Nghe")
            window.update()
            audio = r.listen(source, timeout=4, phrase_time_limit=4)
        try:
            sentence = r.recognize_google(audio, language="vi-VI")
        except:
            sentence = "Hmn"

        print("You: " + sentence)
        Yousen = f"{sentence}"
        var_user.set("You: " + Yousen)
        window.update()

        translator = Translator()
        translated = translator.translate(sentence, src='vi', dest='en')
        # tran=translated.text
        # print("tra", translated.text)
        print("y", translated.text)
        sentenc1 = tokenize(translated.text)
        X = bag_of_words(sentenc1, all_words)
        X = X.reshape(1, X.shape[0])
        X = torch.from_numpy(X).to(device)

        output = model(X)
        _, predicted = torch.max(output, dim=1)
        tag = tags[predicted.item()]
        probs = torch.softmax(output, dim=1)
        prob = probs[0][predicted.item()]
        print("po", probs)
        if prob.item() > 0.85:
            for intent in intents['intents']:
                if tag == intent["tag"]:

                    text = random.choice(intent['responses'])
                    translator = Translator()
                    translated = translator.translate(text, src='en', dest='vi')
                    fb = translated.text
                    print(bot_name, ":" + fb)

                    var_bot.set(bot_name+":"+fb)
                    window.update()

                    output = gTTS(fb, lang="vi", slow=False)
                    output.save("output.mp3")
                    playsound.playsound('output.mp3', True)
                    os.remove("output.mp3")

            if tag == "goodbye":
                detect_red()
                print("hh")
                break
            elif tag == "send faculty Mechatronic":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("cdt")
                open_CDT()
                break
            elif tag == "send faculty Automation":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("tdh")
                open_TDH()
                break
            elif tag == "laguane En":
                for i in range(len(bt)):
                    bt.pop()
                    # print("i", i)
                bt.append("en")
                chatbot_en()
                print("VN")
                break
            elif tag == "open youtube":
                webbrowser.open("youtube.com")
                var_bot.set("Vâng")
                window.update()
                break
            elif tag == "Introduction Faculty":
                # webbrowser.open("https://www.youtube.com/watch?v=KjiNHl-LTOc")
                var_bot.set("Xem video để biết thêm")
                window.update()
                break
            elif tag == "University introduction":
                # webbrowser.open("https://www.youtube.com/watch?v=iEoO2Jw59vs&t=419s")
                var_bot.set("Xem video để biết thêm")
                window.update()
                break
        else:
            Err = f"Err{sentence}"
            print("E:", Err)
            if Err == "ErrHmn":
                dem1 = dem1 + 1
                var_user.set("You: Hmm")
                var_bot.set(bot_name + "Hmm")
                window.update()

            if Err != "ErrHmn":
                dem1 = 0
                print("luu data")
                wb = openpyxl.load_workbook('Data_Management.xlsx')
                sh2 = wb['Data_Err']
                row2 = sh2.max_row + 1
                now = datetime.now()
                stt = row2 - 1
                sh2.cell(row=row2, column=1, value=stt)
                sh2.cell(row=row2, column=2, value=now.strftime("%d"))
                sh2.cell(row=row2, column=3, value="Tháng " + now.strftime("%m"))
                sh2.cell(row=row2, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
                sh2.cell(row=row2, column=5, value=Err)
                wb.save("Data_Management.xlsx")

                var_bot.set(bot_name + " :Tôi không hiểu bạn có thể nói lại được không")
                window.update()

                output = gTTS("Tôi Không hiểu bạn có thể hỏi một số câu như sau", lang="vi", slow=False)
                output.save("output.mp3")
                playsound.playsound('output.mp3', True)
                os.remove("output.mp3")
                time.sleep(0.5)

                chd = []
                for intent in intents['intents']:
                    chd.append(f"-{random.choice(intent['patterns'])}")
                # cauhd=(chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)],"\n",chd[random.randint(1, 10)])
                # cauhd = ["Ask some questions like:", "\n", chd[0], "\n", chd[1], "\n", chd[2], "\n", chd[3],
                #          "\n", chd[4], "\n", chd[5], "\n", chd[6], "\n", chd[7], "\n", chd[8], "\n", chd[9],
                #          "\n", chd[10]]
                var_user.set("-")
                var_bot.set("Trưởng khoa Cơ khí là ai\nTôi muốn gửi một tài liệu tới Automation\nTôi muốn gửi một tài liệu đến cơ điện tử\nSố điện thoại của Kiên là gì?")
                window.update()
                time.sleep(7)
                print(f"{bot_name} : tôi không hiểu...")

        if dem1 > 5:
            var_bot.set("Tạm biệt hẹn gặp lại")
            var_user.set("----")
            window.update()

            output = gTTS("Tạm biệt hẹn gặp lại", lang="vi", slow=False)
            output.save("output.mp3")
            playsound.playsound('output.mp3', True)
            os.remove("output.mp3")
            time.sleep(0.5)

            detect_red()
            break

def setting():
    u1 = "Admin"
    p1 = "123"
    u2 = "Kien"
    p2 = "321"

    def lg_vn():
        print("i1",lg)
        for i in range(len(lg)):
            lg.pop()
            print("i",i)

        lg.append("vn")
        Pw.destroy()
        chatbot_vn()
    def lg_en():
        for i in range(len(lg)):
            lg.pop()
            print("i",i)

        lg.append("en")
        Pw.destroy()
        chatbot_en()

    def login_Ex():
        uname = e1.get()
        password = e2.get()
        if (uname == "" and password == ""):
            messagebox.showinfo("", "You have not entered account")
            Pw.destroy()
            chatbot_en()
        elif (uname == u2 and password == p2):
            path = 'D:\login_picture\Admin_access_to_system'
            now = datetime.now()
            d = now.strftime("%d_%m_%Y_%H_%M_%S")
            frame = cam.read()
            filename = "Acc-"+u2+"-" + d + ".jpg"
            cv2.imwrite(os.path.join(path, filename), frame)

            link = 'HYPERLINK' + '(' + '"' + 'D:\login_picture\Admin_access_to_system\Acc-'+u2+"-" + d + ".jpg"+'")'
            # print("l", link)



            wb = openpyxl.load_workbook('Data_Management.xlsx')
            sh1 = wb['Data_login']
            row = sh1.max_row + 1
            stt = row - 1
            sh1.cell(row=row, column=1, value=stt)
            sh1.cell(row=row, column=2, value=now.strftime("%d"))
            sh1.cell(row=row, column=3, value="Tháng " + now.strftime("%m"))
            sh1.cell(row=row, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
            sh1.cell(row=row, column=5, value=link)
            sh1.cell(row=row, column=6, value="Excel")
            wb.save("Data_Management.xlsx")

            messagebox.showinfo("", "Login Sucess")
            Pw.destroy()
            path = 'Dashboard_Analytics.xlsm'
            os.system(path)
        else:
            messagebox.showinfo("", "Incorrect Username or Password")
            Pw.destroy()
            chatbot_en()

    def login_Box():
        uname = e1.get()
        password = e2.get()

        if (uname == "" and password == ""):
            messagebox.showinfo("", "You have not entered account")
            Pw.destroy()
            chatbot_en()

        elif (uname == u1 and password == p1):
            for i in range(len(quataiCDT)):
                quataiCDT.pop()
                # print("i", i)
            for i in range(len(quataiTDH)):
                quataiTDH.pop()
                # print("i", i)
            path = 'D:\login_picture\Admin_access_to_system'
            now = datetime.now()
            d = now.strftime("%d_%m_%Y_%H_%M_%S")
            frame = cam.read()
            filename = "Acc-"+u1+"-" + d + ".jpg"
            cv2.imwrite(os.path.join(path, filename), frame)

            link = 'HYPERLINK' + '(' + '"' + 'D:\login_picture\Admin_access_to_system\Acc-'+u1+"-" + d + ".jpg"+'")'
            # print("l", link)


            wb = openpyxl.load_workbook('Data_Management.xlsx')
            sh1 = wb['Data_login']
            row = sh1.max_row + 1
            stt = row - 1
            sh1.cell(row=row, column=1, value=stt)
            sh1.cell(row=row, column=2, value=now.strftime("%d"))
            sh1.cell(row=row, column=3, value="Tháng " + now.strftime("%m"))
            sh1.cell(row=row, column=4, value=now.strftime("%d/%m/%Y %H:%M"))
            sh1.cell(row=row, column=5, value=link)
            sh1.cell(row=row, column=6, value="Box")
            wb.save("Data_Management.xlsx")
            door()
            time.sleep(1)
            # door()
            messagebox.showinfo("", "Login Sucess")
            Pw.destroy()
        else:
            messagebox.showinfo("", "Incorrect Username or Password")
            Pw.destroy()
            chatbot_en()

    Pw = Toplevel()
    Pw.iconbitmap('photo_logo.ico')
    Pw.title("Login")
    Pw.geometry("700x250")
    # 560x230
    global e1
    global e2
    Label(Pw, text="Log in", font=("UTM Avo", 15)).place(x=170, y=10)
    Label(Pw, text="UserName", font=("UTM Avo", 15)).place(x=10, y=80)
    Label(Pw, text="Password", font=("UTM Avo", 15)).place(x=10, y=120)

    e1 = Entry(Pw)
    e1.place(x=120, y=90)

    e2 = Entry(Pw)
    e2.place(x=120, y=130)
    e2.config(show="*")

    # Button(Pw,image=photo_b1, text="Login", command=cl , heigh=3, width=13).place(x=155, y=180)

    btn8 = Button(Pw,
                  image=photo_b9,
                  command=lg_vn,
                  border=0, )
    # btn8.pack(side=CENTER)
    btn8.pack(side=RIGHT)

    btn9 = Button(Pw,
                  image=photo_b10,
                  command=lg_en,
                  border=0, )
    # btn9.pack(side=CENTER)
    btn9.pack(side=RIGHT)

    btn6 = Button(Pw,
                  image=photo_b6,
                  command=login_Ex,
                  border=0, )
    btn6.pack(side=RIGHT)

    btn7 = Button(Pw,
                  image=photo_b7,
                  command=login_Box,
                  border=0, )
    btn7.pack(side=RIGHT)

    Pw.mainloop()

def update(ind):# update lable gui
    frame = frames[(ind)%100]
    ind += 1
    label.configure(image=frame)
    window.after(100, update, ind)

#########Def Phuoc
def face_detection5s():##folow face
    dem6=0
    for i in range(len(bt)):
        bt.pop()
        # print("i", i)
    bt.append("none")
    # Load Tensorflow model
    detection_graph = tf.Graph()
    with detection_graph.as_default():
        od_graph_def = tf.GraphDef()
        with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
            serialized_graph = fid.read()
            od_graph_def.ParseFromString(serialized_graph)
            tf.import_graph_def(od_graph_def, name='')

        sess = tf.Session(graph=detection_graph)

    image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')

    # Each box represents a part of the image where a particular object was detected.
    detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')

    # Each score represent how level of confidence for each of the objects.
    # Score is shown on the result image, together with the class label.
    detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')

    # Actual detection.
    detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')
    num_detections = detection_graph.get_tensor_by_name('num_detections:0')

    # Start video stream
    # cap = WebcamVideoStream(0).start()
    # fps = FPS().start()

    while True:

        frame = cam.read()
        # time.sleep(0.1)
        # khung hinh 480x640
        frame = cv2.flip(frame, 1)
        #  dầu vào mô hình thường 300x300 theo thứ tự GRB
        # frame = cv2.resize(frame,(160,140))

        # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
        expanded_frame = np.expand_dims(frame, axis=0)
        (boxes, scores, classes, num_c) = sess.run(
            [detection_boxes, detection_scores, detection_classes, num_detections],
            feed_dict={image_tensor: expanded_frame})
        # Visualization of the detection
        coords = vis_util.visualize_boxes_and_labels_on_image_array(
            frame,
            np.squeeze(boxes),
            np.squeeze(classes).astype(np.int32),
            np.squeeze(scores),
            category_index,
            use_normalized_coordinates=True,
            line_thickness=2,
            min_score_thresh=0.4)
        # print(coords)
        try:
            if (coords != None):
                for coord in coords:
                    # q = q+1
                    # print('q:', q)
                    # if q == 5 :
                    (y1, y2, x1, x2, accuracy, classificaion) = coord
                    w = x2 - x1
                    h = y2 - y1
                    Cir = x1 + ((w) / 2)
                    # print(Cir)
                    cv2.rectangle(frame, (x1, y1), (x1 + w, y1 + h), (0, 255, 255), 1)
                    if not coord:
                        # print(coord)
                        # print("abc")
                        print("looking for questure")
                    else:
                        # print(coord)
                        S = w * h
                    # print(S)
            else:
                S = 1
            print(S)
        except:
            pass
        if (S < 10):
            found = 0
        else:
            found = 1

        if (found == 0):
            print('Stop')
            dem6=0
            # Stop()
        #
        elif (found == 1):
            dem6=dem6+1
            print("dem5s",dem6)
            if dem6==40:
                dem6=0
                # fps.stop()
                face_detection()
                # break

        # cv2.imshow('Detection', frame)
        # fps.update()

        # if cv2.waitKey(1) == ord('q'):
        #     fps.stop()
        #     break

    # print("Fps: {:.2f}".format(fps.fps()))
    # fps.update()
    # cam.stop()
    # cv2.destroyAllWindows()

def face_detection():##folow face
    dem5=0
    for i in range(len(bt)):
        bt.pop()
        # print("i", i)
    bt.append("none")
    # Load Tensorflow model
    detection_graph = tf.Graph()
    with detection_graph.as_default():
        od_graph_def = tf.GraphDef()
        with tf.gfile.GFile(PATH_TO_CKPT, 'rb') as fid:
            serialized_graph = fid.read()
            od_graph_def.ParseFromString(serialized_graph)
            tf.import_graph_def(od_graph_def, name='')

        sess = tf.Session(graph=detection_graph)

    image_tensor = detection_graph.get_tensor_by_name('image_tensor:0')

    # Each box represents a part of the image where a particular object was detected.
    detection_boxes = detection_graph.get_tensor_by_name('detection_boxes:0')

    # Each score represent how level of confidence for each of the objects.
    # Score is shown on the result image, together with the class label.
    detection_scores = detection_graph.get_tensor_by_name('detection_scores:0')

    # Actual detection.
    detection_classes = detection_graph.get_tensor_by_name('detection_classes:0')
    num_detections = detection_graph.get_tensor_by_name('num_detections:0')

    # Start video stream
    # cap = WebcamVideoStream(0).start()
    # fps = FPS().start()
    initial = 10
    while True:

        frame = cam.read()
        # time.sleep(0.1)
        # khung hinh 480x640
        frame = cv2.flip(frame, 1)
        #  dầu vào mô hình thường 300x300 theo thứ tự GRB
        # frame = cv2.resize(frame,(160,140))

        # Expand dimensions since the model expects images to have shape: [1, None, None, 3]
        expanded_frame = np.expand_dims(frame, axis=0)
        (boxes, scores, classes, num_c) = sess.run(
            [detection_boxes, detection_scores, detection_classes, num_detections],
            feed_dict={image_tensor: expanded_frame})
        # Visualization of the detection
        coords = vis_util.visualize_boxes_and_labels_on_image_array(
            frame,
            np.squeeze(boxes),
            np.squeeze(classes).astype(np.int32),
            np.squeeze(scores),
            category_index,
            use_normalized_coordinates=True,
            line_thickness=2,
            min_score_thresh=0.4)
        # print(coords)
        try:
            if (coords != None):
                for coord in coords:
                    # q = q+1
                    # print('q:', q)
                    # if q == 5 :
                    (y1, y2, x1, x2, accuracy, classificaion) = coord
                    w = x2 - x1
                    h = y2 - y1
                    Cir = x1 + ((w) / 2)
                    # print(Cir)
                    cv2.rectangle(frame, (x1, y1), (x1 + w, y1 + h), (0, 255, 255), 1)
                    if not coord:
                        # print(coord)
                        # print("abc")
                        print("looking for questure")
                    else:
                        # print(coord)
                        S = w * h
                    # print(S)
            else:
                S = 1
            print(S)
        except:
            pass
        if (S < 10):
            found = 0
        else:
            found = 1
        flag = 0
        if (found == 0):
            print('Stop')
            dem5=0
            Stop()
        #
        elif (found == 1):

            if (Cir < (0.3 * 640)):
                print('Turn Left')
                dem5=0
                Turnleft()
            elif (Cir > (640 - (0.3 * 640))):
                print('Turn Right')
                dem5=0
                Turnright()
            # elif (S < 9000):
            #     Forward()
            elif (S > initial):
                initial2 = 9000
                if (S < initial2):
                    print("forward")
                    dem5=0
                    Forward()
                elif (S > 40000):
                    print("Reverse")
                    dem5=0
                    Reverse()
                elif (S<40000):
                    dem5=dem5+1
                    print("dem5",dem5)
                if dem5==30:
                    fps.stop()
                    print("chuỷen")
                    Stop()
                    chatbot_en()
                    break

                # elif (S > 40000):
                # Stop()
            # q = q + 1
            # if (S > 40000):
            #     print('Reverse')
            #     q=0
            #     Reverse()
            #
            # elif (S < 40000):
            #     print('Stop')
            #     Stop()
            # # if min_score_thresh >= 90:
            # if (S < 10000):
            #     q=q+1
            #     #  quét khuono mặt trong 20s thì mới bám
            #     print("Q",q)
            #     if q >= 30:
            #         print('Forward')
            #         print('q:', q )
            #     Forward()

        # cv2.rectangle(frame,((640),(480)),(255,255,255),3)
        # cv2.imshow('Detection', frame)
        fps.update()

        if cv2.waitKey(1) == ord('q'):
            fps.stop()
            break

    print("Fps: {:.2f}".format(fps.fps()))
    fps.update()
    cam.stop()
    cv2.destroyAllWindows()

def segment_colour(frame):  #folow red (parking)
    hsv_roi = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    mask_1 = cv2.inRange(hsv_roi, np.array([160, 160, 10]), np.array([190, 255, 255]))
    ycr_roi = cv2.cvtColor(frame, cv2.COLOR_BGR2YCrCb)
    mask_2 = cv2.inRange(ycr_roi, np.array((0., 165., 0.)), np.array((255., 255., 255.)))

    mask = mask_1 | mask_2
    kern_dilate = np.ones((8, 8), np.uint8)
    kern_erode = np.ones((3, 3), np.uint8)
    mask = cv2.erode(mask, kern_erode)  # Eroding
    mask = cv2.dilate(mask, kern_dilate)  # Dilating
    # cv2.imshow('mask',mask)
    return mask

def find_blob(blob):  # returns the red colored circle
    largest_contour = 0
    cont_index = 0
    contours, hierarchy = cv2.findContours(blob, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    for idx, contour in enumerate(contours):
        area = cv2.contourArea(contour)
        if (area > largest_contour):
            largest_contour = area

            cont_index = idx
            # if res>15 and res<18:
            #    cont_index=idx

    r = (0, 0, 2, 2)
    if len(contours) > 0:
        r = cv2.boundingRect(contours[cont_index])

    return r, largest_contour

def target_hist(frame):
    hsv_img = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

    hist = cv2.calcHist([hsv_img], [0], None, [50], [0, 255])
    return hist

def rotate():
    dem2=0
    time.sleep(1)
    while True:
        dem2=dem2+1
        time.sleep(0.2)
        Turnright()
        print("dem",dem2)
        if dem2==32:
            time.sleep(1)
            Stop()
            print("st")
            face_detection5s()
            break

def detect_red():
    i = 0
    quay = 0
    # capture frames from the camera
    while True:
        # grab the raw NumPy array representing the image, then initialize the timestamp and occupied/unoccupied text
        frame = cam.read()
        # // thời gian chò tin hiẹu gui xuong
        # time.sleep(0.1)
        # // lật ảnh
        frame = cv2.flip(frame, (1))

        # if not ret:
        # break
        global centre_x
        global centre_y
        centre_x = 0.
        centre_y = 0.
        hsv1 = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        mask_red = segment_colour(frame)  # masking red the frame
        loct, area = find_blob(mask_red)
        x, y, w, h = loct
        if (w * h) < 200:
            found = 0
        else:
            found = 1
            simg2 = cv2.rectangle(frame, (x, y), (x + w, y + h), 255, 2)
            centre_x = x + ((w) / 2)
            centre_y = y + ((h) / 2)
            cv2.circle(frame, (int(centre_x), int(centre_y)), 3, (0, 110, 255), -1)

        initial = 200
        S = w * h
        # print('s = ',S)
        flag = 0
        #  không phát hiện đối tượng trong khung hình
        if (found == 0):
            # found Red
            i = i + 1
            # #  quay  một lúc nêú không thấy, nó sẽ dừng 50s rồi tiếp tục quét
            if (i >= 20):
                print('Stop')
                Stop()
                if i >= 40:
                    print('turn right')
                    Turnright()
                    if i >= 60:
                        print('STOP')
                        Stop()
                        i = 0

        #  khi phát hiện đối tượng
        elif (found == 1):
           # ưu tiên rẽ trái phải
            i = 0
            if (centre_x <= (0.3 * 640)):
                print('turn Left')
                quay = 0
                Turnleft()
            elif (centre_x >= (640 - (640 * 0.3))):
                print('Turn right')
                quay = 0
                Turnright()
            #  khoảng cách tương đối so với đối tượng được phát hiện (phát hiện >= 5m, Stop <= 1m)
            elif (area > initial):
                initial2 = 50000
                if (area < initial2):
                    print('Forward')
                    quay = 0
                    Forward()
                if (area >= initial2):
                    quay = quay + 1
                    print(quay)
                    # print(' STOP 1 ')
                    if quay > 4:
                        print(' quaybq42r3rb ')
                        Stop()
                        fps.stop()
                        rotate()
                        break


        # cv2.rectangle(frame,(640),(480),(255,255,255),3)
        # cv2.imshow("Original", frame)
        fps.update()

        #      thoi gian cho lay anh
        1000 / 41
        1000 / 1

        if (cv2.waitKey(55) & 0xff == ord('q')):
            fps.stop()
            break
    print("Fps: {:.2f}".format(fps.fps()))
    # fps.update()
    # cam.stop()
    # cv2.destroyAllWindows()

#Tao Gui00
window = Tk()

global var_user #gia tri hoi YOU
global var_bot #gia tri tra loi ROBOT

var_user = StringVar()
var_bot = StringVar()

window.iconbitmap('photo_logo.ico')
window.geometry("1920x1080")

label2 = Label(window, textvariable = var_bot, bg = '#FAB60C')
label2.config(font=("UTM Avo", 20))
var_bot.set('Press the "RUN" button')
label2.pack()

label1 = Label(window, textvariable = var_user, bg = '#ADD8E6')
label1.config(font=("UTM Avo", 20))
var_user.set('Welcome')
label1.pack()

frames = [PhotoImage(file='photo_test.gif',format = 'gif -index %i' %(i)) for i in range(100)]
window.title('KP-CTUT')

label = Label(window, width = 1100, height = 650)
#width = 1100, height = 650(ngang)
#width = 800, height = 1050(doc)

label.pack()
window.after(0, update, 0)


photo_b1 = PhotoImage(file="photo_b1.png") #RUN
photo_b2 = PhotoImage(file="photo_b2.png") #EXIT
photo_b3 = PhotoImage(file="photo_b3.png") #OPEN_A
photo_b4 = PhotoImage(file="photo_b4.png") #OPEN_B
photo_b5 = PhotoImage(file="photo_b5.png") #OPEN_EXCEL+BOX
photo_b6 = PhotoImage(file="photo_b6.png") #Sign in EX
photo_b7 = PhotoImage(file="photo_b7.png") #Sign in BOX
photo_b8 = PhotoImage(file="photo_b8.png") #back in chatbot
photo_b9 = PhotoImage(file="photo_b9.png") #viet nam
photo_b10 = PhotoImage(file="photo_b10.png") #english



btn1 = Button(window,
              image=photo_b1,
              command=face_detection5s,
              border=0,)
btn1.pack(side=LEFT)

btn2 = Button(window,
              image=photo_b2,
              command=window.destroy,
              border=0,)
btn2.pack(side=LEFT)

btn3 = Button(window,
              image=photo_b3,
              command=open_CDT,
              border=0,)
btn3.pack(side=RIGHT)

btn4 = Button(window,
              image=photo_b4,
              command=open_TDH,
              border=0,)
btn4.pack(side=RIGHT)
btn5 = Button(window,
              image=photo_b5,
              command=setting,
              border=0,)
btn5.pack(side=RIGHT)


window.mainloop()




